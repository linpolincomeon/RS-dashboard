#!/usr/bin/env python3
"""Productividad de choferes -> choferes-data.json

Fuente: Google Sheet "Bodegas_por_Camion_v2" (hoja Movimientos), la misma que
alimenta reporte_chofer_conectado.html. NO pasa por Odoo.

Refresh manual:
    1) Descargar el Sheet como .xlsx
       (https://docs.google.com/spreadsheets/d/11UTBZyBicW3HJcwnPpxyK_v_sqz1rPqd5oz17N471hE)
    2) python3 extract_choferes.py Bodegas_por_Camion_v2.xlsx

Métrica central: la jornada de cada camión se parte en tres tramos
    arranque = INICIO -> 1ra VENTA      (camión en marcha, sin vender)
    activo   = 1ra VENTA -> última VENTA
    cola     = última VENTA -> CIERRE   (camión en marcha, sin vender)
"""
import sys, json, datetime as dt, statistics as st
from collections import defaultdict

import openpyxl

VENTANA_DIAS = 56           # ventana de análisis hacia atrás desde el último dato
MIN_VENTAS_DIA = 2          # días con 1 sola venta no permiten medir jornada

# hoja "conductores y sus zonas" del mismo Sheet
CHOFERES = {
    'TJ': ('TJVS-53', 'José Luis Valenzuela', 'Curicó', 'José Luis'),
    'VD': ('VDKT-95', 'Nino Aguilera / F. Garroz', 'San Fernando · VI Costa', 'Nino'),
    'TY': ('TYDG-23', 'Patricio Garrido', 'Linares al sur', 'Patricio'),
    'PY': ('PYHK-28', 'Mario Marín', 'Linares', 'Mario'),
    'SH': ('SHGP-60', 'Roberto Urrutia', 'Linares', 'Roberto'),
    'PH': ('PHXC-44', 'Jorge Aguilera "Tato"', 'Rancagua · Mostazal', 'Tato'),
    'HH': ('HHPT-71', 'Sin conductor (spare)', 'San Fernando', 'HH (spare)'),
}
ORDEN = ['TJ', 'VD', 'TY', 'PY', 'SH', 'PH', 'HH']


def mins(h):
    return h.hour * 60 + h.minute


def leer(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Movimientos' not in wb.sheetnames:
        sys.exit('ERROR: el archivo no tiene hoja "Movimientos"')
    out = []
    for r in wb['Movimientos'].iter_rows(min_row=2, values_only=True):
        f, h, cam, tipo, lts = r[0], r[1], r[2], r[3], r[4]
        if not f or not cam or not tipo or isinstance(f, str):
            continue
        if not isinstance(h, (dt.time, dt.datetime)):
            continue
        if isinstance(h, dt.datetime):
            h = h.time()
        # el formato viejo traía "Nino VD" en la columna Camión
        cam = str(cam).strip().split()[-1].upper()
        if cam not in CHOFERES:
            continue
        try:
            lts = float(lts) if lts not in (None, '') else 0.0
        except (TypeError, ValueError):
            lts = 0.0
        out.append((f.date(), mins(h), cam, str(tipo).strip().upper(), lts))
    return out


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'Bodegas_por_Camion_v2.xlsx'
    mov = leer(path)
    if not mov:
        sys.exit('ERROR: no se leyó ningún movimiento')

    hasta = max(m[0] for m in mov)
    desde = hasta - dt.timedelta(days=VENTANA_DIAS)
    mov = [m for m in mov if m[0] >= desde]

    habiles = {desde + dt.timedelta(days=i) for i in range((hasta - desde).days + 1)
               if (desde + dt.timedelta(days=i)).weekday() < 5}

    byday = defaultdict(list)
    for d, t, c, tipo, l in mov:
        byday[(c, d)].append((t, tipo, l))

    choferes, diario = [], []
    for c in ORDEN:
        dias = []
        for (cc, d), evs in sorted(byday.items()):
            if cc != c:
                continue
            evs.sort()
            ventas = [e for e in evs if e[1] == 'VENTA']
            if not ventas:
                continue
            ini = [e for e in evs if e[1] == 'INICIO']
            cie = [e for e in evs if e[1] == 'CIERRE']
            t0 = min(ini[0][0] if ini else evs[0][0], evs[0][0])
            t1 = max(cie[-1][0] if cie else evs[-1][0], evs[-1][0])
            tv = [e[0] for e in ventas]
            # ventas cargadas a la planilla en bloque (varias en el mismo minuto)
            # => las horas de ese día no reflejan la operación real
            seguidas = sum(1 for i in range(1, len(tv)) if tv[i] - tv[i - 1] <= 2)
            reg = dict(d=d.isoformat(), ini=t0, fin=t1, nv=len(ventas),
                       lts=sum(e[2] for e in ventas), v0=tv[0], v1=tv[-1],
                       batch=seguidas / max(len(tv) - 1, 1))
            dias.append(reg)
            diario.append(dict(cam=c, **reg))

        if not dias:
            continue
        med = [x for x in dias if x['nv'] >= MIN_VENTAS_DIA and x['fin'] > x['ini']]
        batch = st.mean([x['batch'] for x in dias])
        confiable = batch < 0.30 and len(med) >= 5
        fila = dict(
            cam=c, patente=CHOFERES[c][0], chofer=CHOFERES[c][1], zona=CHOFERES[c][2],
            corto=CHOFERES[c][3],
            dias=len(dias),
            dias_habiles=len({dt.date.fromisoformat(x['d']) for x in dias} & habiles),
            ventas_dia=round(st.mean([x['nv'] for x in dias]), 1),
            litros_dia=round(st.mean([x['lts'] for x in dias])),
            litros_total=round(sum(x['lts'] for x in dias)),
            entrega_med=round(st.median([x['lts'] / x['nv'] for x in dias])),
            batch_pct=round(batch, 2), confiable=confiable,
        )
        if med:
            jor = st.mean([(x['fin'] - x['ini']) / 60 for x in med])
            fila.update(
                inicio=round(st.median([x['ini'] for x in med])),
                fin=round(st.median([x['fin'] for x in med])),
                primera_venta=round(st.median([x['v0'] for x in med])),
                ultima_venta=round(st.median([x['v1'] for x in med])),
                arranque=round(st.median([x['v0'] - x['ini'] for x in med])),
                cola=round(st.median([x['fin'] - x['v1'] for x in med])),
                jornada_h=round(jor, 1),
                activo_pct=round(st.mean([(x['v1'] - x['v0']) / (x['fin'] - x['ini']) for x in med]), 2),
                litros_hora=round(st.mean([x['lts'] for x in med]) / jor) if jor else 0,
            )
        choferes.append(fila)

    out = dict(
        generated_utc=dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M+00:00'),
        fuente='Google Sheet Bodegas_por_Camion_v2 · hoja Movimientos',
        desde=desde.isoformat(), hasta=hasta.isoformat(),
        dias_habiles=len(habiles), choferes=choferes, diario=diario,
    )

    # validación: si el dato viene implausible, no escribir
    ok = [c for c in out['choferes'] if c['confiable']]
    if len(ok) < 3:
        sys.exit(f'ERROR: solo {len(ok)} camiones con horas confiables — no se escribe el JSON')
    if sum(c['litros_total'] for c in out['choferes']) < 500_000:
        sys.exit('ERROR: litros totales de la ventana implausibles — no se escribe el JSON')

    with open('choferes-data.json', 'w', encoding='utf-8') as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)

    print(f"choferes-data.json  {out['desde']} -> {out['hasta']}  ({out['dias_habiles']} días hábiles)")
    for c in out['choferes']:
        marca = '' if c['confiable'] else '  ⚠ horas no confiables (carga en bloque)'
        print(f"  {c['cam']:3} {c['chofer'][:26]:26} {c['dias']:3}d  "
              f"{c['litros_dia']:6,} L/día  {c['ventas_dia']:4} vtas/día"
              f"  activo {int(100 * c.get('activo_pct', 0)):3}%{marca}")


if __name__ == '__main__':
    main()
